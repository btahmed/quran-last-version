#!/usr/bin/env python3
"""
Créer un fichier Excel propre avec tous les credentials
"""

import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def creer_excel_credentials():
    """Créer un fichier Excel à partir du CSV"""
    
    # Lire le CSV
    csv_path = Path(__file__).parent / "output" / "nouveaux_credentials_tous.csv"
    
    if not csv_path.exists():
        print(f"❌ Fichier CSV non trouvé: {csv_path}")
        return
    
    print(f"📖 Lecture du fichier CSV...")
    
    # Lire les données du CSV
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        data = list(reader)
    
    print(f"✅ {len(data)} comptes trouvés")
    
    # Créer le fichier Excel
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = Path(__file__).parent / "output" / f"credentials_complets_{timestamp}.xlsx"
    
    print(f"📝 Création du fichier Excel...")
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tous les comptes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    password_font = Font(bold=True, name="Courier New")
    password_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ['Username', 'Password', 'Prénom', 'Nom', 'Rôle', 'Classe']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    for row_num, row_data in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=row_data['Username']).alignment = cell_alignment
        
        # Password avec style spécial
        pwd_cell = ws.cell(row=row_num, column=2, value=row_data['Password'])
        pwd_cell.font = password_font
        pwd_cell.alignment = password_alignment
        pwd_cell.border = border
        
        ws.cell(row=row_num, column=3, value=row_data['Prénom']).alignment = cell_alignment
        ws.cell(row=row_num, column=4, value=row_data['Nom']).alignment = cell_alignment
        ws.cell(row=row_num, column=5, value=row_data['Rôle']).alignment = cell_alignment
        ws.cell(row=row_num, column=6, value=row_data['Classe']).alignment = cell_alignment
        
        # Bordures
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = border
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Sauvegarder
    wb.save(excel_path)
    print(f"✅ Fichier Excel créé: {excel_path}")
    
    # Créer aussi des feuilles séparées par type
    excel_path_separe = Path(__file__).parent / "output" / f"credentials_par_type_{timestamp}.xlsx"
    
    print(f"📝 Création du fichier Excel avec feuilles séparées...")
    
    wb2 = Workbook()
    wb2.remove(wb2.active)  # Supprimer la feuille par défaut
    
    # Fonction pour créer une feuille
    def create_sheet(workbook, sheet_name, data_subset):
        ws = workbook.create_sheet(sheet_name)
        
        # En-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Données
        for row_num, row_data in enumerate(data_subset, 2):
            ws.cell(row=row_num, column=1, value=row_data['Username']).alignment = cell_alignment
            
            pwd_cell = ws.cell(row=row_num, column=2, value=row_data['Password'])
            pwd_cell.font = password_font
            pwd_cell.alignment = password_alignment
            pwd_cell.border = border
            
            ws.cell(row=row_num, column=3, value=row_data['Prénom']).alignment = cell_alignment
            ws.cell(row=row_num, column=4, value=row_data['Nom']).alignment = cell_alignment
            ws.cell(row=row_num, column=5, value=row_data['Rôle']).alignment = cell_alignment
            ws.cell(row=row_num, column=6, value=row_data['Classe']).alignment = cell_alignment
            
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = border
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        
        ws.freeze_panes = 'A2'
    
    # Tous les comptes
    create_sheet(wb2, 'Tous', data)
    
    # Étudiants 8h45
    data_8h45 = [d for d in data if '8h45' in d['Classe'] and d['Rôle'] == 'student']
    create_sheet(wb2, 'Étudiants 8h45', data_8h45)
    
    # Étudiants 10h45
    data_10h45 = [d for d in data if '10h45' in d['Classe'] and d['Rôle'] == 'student']
    create_sheet(wb2, 'Étudiants 10h45', data_10h45)
    
    # Professeurs
    data_profs = [d for d in data if d['Rôle'] == 'teacher']
    create_sheet(wb2, 'Professeurs', data_profs)
    
    wb2.save(excel_path_separe)
    print(f"✅ Fichier Excel avec feuilles séparées créé: {excel_path_separe}")
    
    # Afficher quelques exemples
    print(f"\n📊 EXEMPLES DE COMPTES:")
    print("="*60)
    
    # salma_aneflous
    salma = next((d for d in data if d['Username'] == 'salma_aneflous'), None)
    if salma:
        print(f"\n👤 salma_aneflous:")
        print(f"   Mot de passe: {salma['Password']}")
        print(f"   Nom: {salma['Prénom']} {salma['Nom']}")
        print(f"   Classe: {salma['Classe']}")
    
    # Un professeur
    prof = next((d for d in data if d['Rôle'] == 'teacher'), None)
    if prof:
        print(f"\n👨‍🏫 {prof['Username']}:")
        print(f"   Mot de passe: {prof['Password']}")
        print(f"   Nom: {prof['Prénom']}")
        print(f"   Classe(s): {prof['Classe']}")
    
    print(f"\n📈 STATISTIQUES:")
    print(f"   Total: {len(data)} comptes")
    print(f"   Étudiants 8h45: {len(data_8h45)}")
    print(f"   Étudiants 10h45: {len(data_10h45)}")
    print(f"   Professeurs: {len(data_profs)}")
    
    print(f"\n✅ FICHIERS CRÉÉS:")
    print(f"   1. {excel_path.name}")
    print(f"   2. {excel_path_separe.name}")
    print(f"\n📂 Emplacement: {excel_path.parent}")


if __name__ == "__main__":
    creer_excel_credentials()
