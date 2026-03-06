#!/usr/bin/env python3
"""
Script pour extraire quelques mots de passe des fichiers d'identifiants
et les utiliser pour tester les permissions par classe
"""

from openpyxl import load_workbook
from pathlib import Path


def extraire_mots_de_passe():
    """Extrait quelques mots de passe pour les tests"""
    
    output_dir = Path(__file__).parent / "output"
    
    # Fichiers d'identifiants
    fichiers = {
        'professeurs': output_dir / "credentials_2026-02-20_10-30-16.excel.xlsx",
        'nouveaux_etudiants': output_dir / "credentials_2026-02-20_10-32-32.excel.xlsx", 
        'etudiants_10h45': output_dir / "credentials_2026-02-20_09-59-27.excel.xlsx"
    }
    
    comptes_test = {}
    
    print("🔐 EXTRACTION DES MOTS DE PASSE POUR TESTS")
    print("="*60)
    
    for type_fichier, fichier_path in fichiers.items():
        if not fichier_path.exists():
            print(f"⚠️  Fichier manquant: {fichier_path}")
            continue
            
        try:
            # Lire le fichier Excel avec openpyxl
            workbook = load_workbook(fichier_path)
            sheet = workbook.active
            
            # Trouver les colonnes Username et Password
            headers = {}
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    if 'username' in str(cell_value).lower():
                        headers['username'] = col
                    elif 'password' in str(cell_value).lower():
                        headers['password'] = col
            
            if 'username' not in headers or 'password' not in headers:
                print(f"⚠️  Colonnes Username/Password non trouvées dans {type_fichier}")
                continue
            
            print(f"\n📁 {type_fichier.upper()}")
            
            # Compter les lignes de données
            nb_comptes = 0
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=headers['username']).value:
                    nb_comptes += 1
            
            print(f"📊 {nb_comptes} comptes dans le fichier")
            
            # Prendre les 2 premiers comptes de chaque fichier
            comptes_ajoutes = 0
            for row in range(2, sheet.max_row + 1):
                if comptes_ajoutes >= 2:
                    break
                    
                username = sheet.cell(row=row, column=headers['username']).value
                password = sheet.cell(row=row, column=headers['password']).value
                
                if username and password:
                    comptes_test[str(username)] = str(password)
                    print(f"✅ {username}: {password}")
                    comptes_ajoutes += 1
                
        except Exception as e:
            print(f"❌ Erreur lecture {type_fichier}: {e}")
    
    return comptes_test


def creer_script_test_avec_mots_de_passe(comptes_test):
    """Crée un script de test avec les vrais mots de passe"""
    
    script_content = f'''#!/usr/bin/env python3
"""
Script de test avec mots de passe réels extraits automatiquement
Généré automatiquement le 2026-02-20
"""

import os
import sys
import django
import requests
from pathlib import Path

# Configuration Django
project_root = Path(__file__).parent.parent / "ancien django" / "MYSITEE" / "MYSITEE"
sys.path.append(str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')

try:
    django.setup()
except Exception as e:
    print(f"❌ Erreur configuration Django: {{e}}")
    sys.exit(1)

from django.contrib.auth import authenticate
from django.contrib.auth.models import Group
from django.db import models
from tasks.models import User


class TesteurPermissionsReel:
    """Testeur avec mots de passe réels"""
    
    def __init__(self):
        self.base_url = "http://127.0.0.1:8000"
        self.api_url = f"{{self.base_url}}/api"
        
        # Mots de passe réels extraits des fichiers
        self.comptes_test = {comptes_test}
    
    def tester_connexion_api(self, username, password):
        """Teste la connexion via l'API"""
        try:
            response = requests.post(
                f"{{self.api_url}}/auth/login/",
                json={{'username': username, 'password': password}},
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return True, data.get('access'), data
            else:
                return False, None, f"Erreur {{response.status_code}}: {{response.text}}"
                
        except Exception as e:
            return False, None, f"Erreur: {{e}}"
    
    def tester_acces_utilisateurs(self, token):
        """Teste l'accès à la liste des utilisateurs avec filtrage"""
        try:
            headers = {{'Authorization': f'Bearer {{token}}'}}
            response = requests.get(
                f"{{self.api_url}}/users/",
                headers=headers,
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return True, data
            else:
                return False, f"Erreur {{response.status_code}}: {{response.text}}"
                
        except Exception as e:
            return False, f"Erreur: {{e}}"
    
    def analyser_filtrage(self, username, users_data):
        """Analyse si le filtrage par classe fonctionne"""
        try:
            user = User.objects.get(username=username)
            user_classes = list(user.groups.filter(
                name__in=['Classe_8h45', 'Classe_10h45']
            ).values_list('name', flat=True))
            
            print(f"👤 {{username}} ({{user.role}})")
            print(f"👥 Classes utilisateur: {{', '.join(user_classes)}}")
            print(f"📊 Voit {{users_data['count']}} utilisateurs")
            
            # Analyser les utilisateurs visibles
            classes_vues = set()
            roles_vus = {{'student': 0, 'teacher': 0}}
            
            for user_data in users_data['users']:
                # Récupérer les classes de cet utilisateur visible
                try:
                    user_obj = User.objects.get(id=user_data['id'])
                    user_obj_classes = list(user_obj.groups.filter(
                        name__in=['Classe_8h45', 'Classe_10h45']
                    ).values_list('name', flat=True))
                    
                    classes_vues.update(user_obj_classes)
                    roles_vus[user_data['role']] += 1
                except:
                    pass
            
            print(f"🔍 Classes visibles dans les résultats: {{', '.join(classes_vues)}}")
            print(f"📈 Rôles vus: {{roles_vus['student']}} étudiants, {{roles_vus['teacher']}} professeurs")
            
            # Vérifier le filtrage
            filtrage_correct = True
            
            # Si l'utilisateur n'a pas de classe, il ne devrait rien voir (sauf superuser)
            if not user_classes and not user.is_superuser:
                if users_data['count'] > 0:
                    print("⚠️  PROBLÈME: Utilisateur sans classe voit des données")
                    filtrage_correct = False
            
            # Si l'utilisateur a des classes, il ne devrait voir que ces classes
            elif user_classes:
                for classe_vue in classes_vues:
                    if classe_vue not in user_classes:
                        print(f"⚠️  PROBLÈME: Voit la classe {{classe_vue}} mais n'y appartient pas")
                        filtrage_correct = False
            
            if filtrage_correct:
                print("✅ Filtrage par classe CORRECT")
            else:
                print("❌ Problème de filtrage détecté")
            
            return filtrage_correct
            
        except Exception as e:
            print(f"❌ Erreur analyse: {{e}}")
            return False
    
    def tester_utilisateur_complet(self, username, password):
        """Test complet d'un utilisateur"""
        print(f"\\n🧪 TEST COMPLET: {{username}}")
        print("-" * 50)
        
        # Test connexion
        success, token, data = self.tester_connexion_api(username, password)
        
        if not success:
            print(f"❌ Connexion échouée: {{data}}")
            return False
        
        print(f"✅ Connexion réussie")
        print(f"👤 Profil API: {{data.get('first_name', 'N/A')}} - {{data.get('role', 'N/A')}}")
        
        # Test accès utilisateurs
        success, users_data = self.tester_acces_utilisateurs(token)
        
        if not success:
            print(f"❌ Accès utilisateurs échoué: {{users_data}}")
            return False
        
        print(f"✅ Accès utilisateurs réussi")
        
        # Analyser le filtrage
        filtrage_ok = self.analyser_filtrage(username, users_data)
        
        return filtrage_ok
    
    def executer_tous_les_tests(self):
        """Exécute tous les tests"""
        print("🧪 TESTS COMPLETS AVEC MOTS DE PASSE RÉELS")
        print("="*60)
        
        resultats = {{}}
        
        for username, password in self.comptes_test.items():
            try:
                # Vérifier que le compte existe
                if not User.objects.filter(username=username).exists():
                    print(f"⚠️  {{username}}: Compte non trouvé, ignoré")
                    continue
                
                # Test complet
                success = self.tester_utilisateur_complet(username, password)
                resultats[username] = success
                
            except Exception as e:
                print(f"❌ Erreur test {{username}}: {{e}}")
                resultats[username] = False
        
        # Statistiques finales
        print("\\n📊 RÉSULTATS FINAUX")
        print("="*40)
        
        reussis = sum(1 for success in resultats.values() if success)
        total = len(resultats)
        
        print(f"✅ Tests réussis: {{reussis}}/{{total}}")
        
        for username, success in resultats.items():
            status = "✅" if success else "❌"
            user = User.objects.get(username=username)
            classes = list(user.groups.filter(
                name__in=['Classe_8h45', 'Classe_10h45']
            ).values_list('name', flat=True))
            classes_str = ', '.join(classes) if classes else 'Aucune'
            
            print(f"   {{status}} {{username}} ({{user.role}}) - {{classes_str}}")
        
        if reussis == total and total > 0:
            print("\\n🎉 TOUS LES TESTS RÉUSSIS!")
            print("✅ Le système de permissions par classe fonctionne parfaitement")
        elif reussis > 0:
            print(f"\\n⚠️  {{total - reussis}} tests ont échoué sur {{total}}")
            print("ℹ️  Vérifiez la configuration pour les comptes en échec")
        else:
            print("\\n❌ TOUS LES TESTS ONT ÉCHOUÉ")
            print("ℹ️  Vérifiez la configuration du middleware et des groupes")
        
        return reussis, total


def main():
    """Point d'entrée principal"""
    testeur = TesteurPermissionsReel()
    reussis, total = testeur.executer_tous_les_tests()
    
    if reussis == total and total > 0:
        print("\\n🎯 SYSTÈME VALIDÉ - Prêt pour la production!")
    else:
        print("\\n🔧 Ajustements nécessaires")


if __name__ == "__main__":
    main()
'''
    
    # Écrire le script
    script_path = Path(__file__).parent / "test_permissions_reel.py"
    with open(script_path, 'w', encoding='utf-8') as f:
        f.write(script_content)
    
    print(f"\n✅ Script créé: {script_path}")
    return script_path


def main():
    """Point d'entrée principal"""
    print("Extraction des mots de passe pour tests...")
    
    # Extraire les mots de passe
    comptes_test = extraire_mots_de_passe()
    
    if not comptes_test:
        print("\n❌ Aucun mot de passe extrait")
        return
    
    print(f"\n📊 {len(comptes_test)} comptes extraits pour les tests")
    
    # Créer le script de test
    script_path = creer_script_test_avec_mots_de_passe(comptes_test)
    
    print(f"\n🚀 PROCHAINE ÉTAPE:")
    print(f"   python {script_path.name}")


if __name__ == "__main__":
    main()