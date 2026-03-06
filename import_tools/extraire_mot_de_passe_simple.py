#!/usr/bin/env python3
"""
Script simple pour extraire le mot de passe de salma_aneflous
"""

import sys
from pathlib import Path


def extraire_mot_de_passe_salma():
    """Extrait le mot de passe de salma_aneflous du fichier credentials"""
    
    # salma_aneflous est dans le premier import (étudiants 10h45)
    fichier_path = Path(__file__).parent / "output" / "credentials_2026-02-20_09-59-27.excel.xlsx"
    
    print(f"🔍 EXTRACTION MOT DE PASSE POUR salma_aneflous")
    print("="*60)
    
    if not fichier_path.exists():
        print(f"❌ Fichier non trouvé: {fichier_path}")
        return None
    
    print(f"📁 Fichier: {fichier_path}")
    
    # Méthode 1: Essayer avec openpyxl si disponible
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(fichier_path)
        sheet = workbook.active
        
        # Trouver les colonnes
        headers = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                if 'username' in str(cell_value).lower():
                    headers['username'] = col
                elif 'password' in str(cell_value).lower():
                    headers['password'] = col
        
        if 'username' not in headers or 'password' not in headers:
            print("❌ Colonnes Username/Password non trouvées")
            return None
        
        # Chercher salma_aneflous
        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row=row, column=headers['username']).value
            if username and str(username).lower() == 'salma_aneflous':
                password = sheet.cell(row=row, column=headers['password']).value
                print(f"✅ TROUVÉ!")
                print(f"👤 Username: {username}")
                print(f"🔐 Password: {password}")
                return str(password)
        
        print("❌ salma_aneflous non trouvé dans le fichier")
        return None
        
    except ImportError:
        print("❌ openpyxl non disponible")
        return None
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return None


def main():
    """Point d'entrée principal"""
    password = extraire_mot_de_passe_salma()
    
    if password:
        print(f"\n🎯 MOT DE PASSE POUR salma_aneflous: {password}")
        print(f"\n📋 MAINTENANT VOUS POUVEZ TESTER:")
        print(f"   Username: salma_aneflous")
        print(f"   Password: {password}")
    else:
        print(f"\n❌ Impossible d'extraire le mot de passe")
        print(f"📁 Ouvrez manuellement: credentials_2026-02-20_09-59-27.excel.xlsx")
        print(f"🔍 Cherchez la ligne avec 'salma_aneflous'")


if __name__ == "__main__":
    main()